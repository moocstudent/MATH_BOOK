"""
装柜报告导入/导出工具
  - parse_items_excel : 从 Excel 解析货物清单
  - build_template    : 生成货物清单模板 (xlsx)
  - build_excel_report: 生成装柜结果报告 (xlsx)
  - build_pdf_report  : 生成装柜结果报告 (pdf, 支持中文)
  - build_html_report : 生成装柜结果报告 (html, 页面查看)
所有尺寸单位 cm、重量 kg；报表中公差显示 mm、体积显示 m³。
"""
from __future__ import annotations

import io
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 货物清单列头(中文 -> 字段)
ITEM_COLUMNS = ["名称", "长(cm)", "宽(cm)", "高(cm)", "重量(kg)", "数量", "可堆叠", "颜色"]
_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_TITLE = Font(bold=True, size=14)


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "货物清单"
    ws.append(ITEM_COLUMNS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    samples = [
        ["PalletBox-A", 120, 100, 110, 260, 18, "是", "#4C72B0"],
        ["PalletBox-B", 110, 90, 90, 180, 20, "是", "#DD8452"],
        ["Fragile-Top", 100, 80, 70, 70, 8, "否", "#DA8BC3"],
    ]
    for row in samples:
        ws.append(row)
    for i, w in enumerate([16, 9, 9, 9, 10, 8, 8, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    return _save(wb)


def parse_items_excel(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    items = []
    for r in rows[1:]:
        if r is None or r[0] in (None, ""):
            continue
        name = str(r[0])
        try:
            length, width, height = float(r[1]), float(r[2]), float(r[3])
            weight = float(r[4])
            qty = int(float(r[5])) if r[5] not in (None, "") else 1
        except (ValueError, TypeError):
            continue
        stack_raw = str(r[6]).strip().lower() if len(r) > 6 and r[6] is not None else "是"
        stackable = stack_raw in ("是", "y", "yes", "true", "1", "t")
        color = str(r[7]) if len(r) > 7 and r[7] not in (None, "") else None
        items.append({
            "name": name, "length": length, "width": width, "height": height,
            "weight": weight, "qty": qty, "stackable": stackable, "color": color,
        })
    return items


# ---------------------------------------------------------------------------
# 报表统计(参考 LoadExpert 报表布局)
# ---------------------------------------------------------------------------

def _container_tare_kg(length_cm: float) -> float:
    """20 尺约 2800kg,40 尺约 5000kg(与参考报表一致)。"""
    return 2800.0 if length_cm < 700 else 5000.0


def _placement_volume_m3(p: dict) -> float:
    return p["dx"] * p["dy"] * p["dz"] / 1_000_000


def _aggregate_placements(placements: list[dict]) -> list[dict]:
    """按货物基础名称汇总:装载数量、小件数、说明。"""
    meta: dict[str, dict] = {}
    counts = Counter()
    for p in placements:
        base = p["base_name"]
        counts[base] += 1
        if base not in meta:
            meta[base] = {
                "desc": p.get("desc", ""),
                "pcs_per_box": max(1, int(p.get("pcs_per_box", 1))),
                "price": float(p.get("price", 0)),
            }
    return [
        {
            "name": name,
            "desc": meta[name]["desc"],
            "qty": cnt,
            "pcs": cnt * meta[name]["pcs_per_box"],
            "price": meta[name]["price"] * cnt,
        }
        for name, cnt in sorted(counts.items())
    ]


def _tolerances_mm(c: dict) -> tuple[float, float, float]:
    L, W, H = c["length"], c["width"], c["height"]
    if not c.get("placements"):
        return L * 10, W * 10, H * 10
    max_x = max(p["x"] + p["dx"] for p in c["placements"])
    max_y = max(p["y"] + p["dy"] for p in c["placements"])
    max_z = max(p["z"] + p["dz"] for p in c["placements"])
    return (L - max_x) * 10, (W - max_y) * 10, (H - max_z) * 10


def _container_type_name(name: str) -> str:
    """去掉多柜后缀 '-2' 等,保留箱型名称。"""
    if "-" in name:
        base, suffix = name.rsplit("-", 1)
        if suffix.isdigit():
            return base
    return name


def build_report_data(result: dict) -> dict:
    containers = [c for c in result.get("containers", []) if c.get("item_count", 0) > 0]
    all_placements = [p for c in containers for p in c.get("placements", [])]

    total_pcs = len(all_placements)
    total_volume = sum(_placement_volume_m3(p) for p in all_placements)
    total_weight = sum(c["used_weight"] for c in containers)
    total_price = sum(
        float(p.get("price", 0))
        for p in all_placements
    )

    container_reports = []
    for idx, c in enumerate(containers, start=1):
        vol = sum(_placement_volume_m3(p) for p in c["placements"])
        tol_l, tol_w, tol_h = _tolerances_mm(c)
        tare = _container_tare_kg(c["length"])
        c_price = sum(float(p.get("price", 0)) for p in c["placements"])
        container_reports.append({
            "index": idx,
            "type_name": _container_type_name(c["name"]),
            "volume_m3": vol,
            "volume_util": c["volume_utilization"],
            "weight_kg": c["used_weight"],
            "gross_kg": round(c["used_weight"] + tare, 2),
            "price": c_price,
            "tol_length_mm": tol_l,
            "tol_width_mm": tol_w,
            "tol_height_mm": tol_h,
            "total_pcs": c["item_count"],
            "items": _aggregate_placements(c["placements"]),
        })

    return {
        "total_containers": len(containers),
        "total_pcs": total_pcs,
        "total_volume_m3": total_volume,
        "total_weight_kg": total_weight,
        "total_price": total_price,
        "items": _aggregate_placements(all_placements),
        "containers": container_reports,
        "unpacked_count": len(result.get("unpacked", [])),
    }


def _fmt2(v: float) -> str:
    return f"{v:.2f}"


def _item_table_rows(items: list[dict]) -> list[list[str]]:
    rows = [["名称", "说明", "装载数量", "小件数"]]
    for it in items:
        rows.append([it["name"], it.get("desc", ""), str(it["qty"]), str(it["pcs"])])
    return rows


def build_html_report(result: dict) -> str:
    d = build_report_data(result)
    parts = [
        '<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">',
        '<title>装柜报表</title>',
        "<style>",
        "body{font-family:'Microsoft YaHei','Segoe UI',sans-serif;margin:24px 32px;color:#111;}",
        "h1{font-size:20px;margin:0 0 20px;}",
        "h2{font-size:15px;margin:24px 0 10px;border-bottom:1px solid #333;padding-bottom:4px;}",
        "h3{font-size:14px;margin:18px 0 8px;}",
        ".kv{margin:4px 0;font-size:13px;}",
        ".kv b{display:inline-block;min-width:160px;}",
        "table{border-collapse:collapse;width:100%;max-width:640px;margin:8px 0 16px;font-size:13px;}",
        "th,td{border:1px solid #333;padding:6px 10px;text-align:left;}",
        "th{background:#f5f5f5;}",
        ".grid{display:grid;grid-template-columns:1fr 1fr;gap:4px 24px;max-width:720px;font-size:13px;}",
        ".sep{border-top:1px solid #999;margin:20px 0;}",
        "</style></head><body>",
        "<h1>报表</h1>",
        "<h2>总体装柜情况</h2>",
        f'<div class="kv"><b>总集装箱数</b> {d["total_containers"]}</div>',
        f'<div class="kv"><b>货物总数(PCS)</b> {d["total_pcs"]}</div>',
        f'<div class="kv"><b>货物总体积(m³)</b> {_fmt2(d["total_volume_m3"])}</div>',
        f'<div class="kv"><b>货物总重量(kg)</b> {_fmt2(d["total_weight_kg"])}</div>',
        f'<div class="kv"><b>货物总价格</b> {_fmt2(d["total_price"])}</div>',
        f'<h3>共装上货物(PCS) {d["total_pcs"]}</h3>',
        _html_table(_item_table_rows(d["items"])),
    ]
    for cr in d["containers"]:
        parts += [
            '<div class="sep"></div>',
            f'<h2>集装箱 {cr["index"]}</h2>',
            f'<div class="kv"><b>集装箱</b> {cr["type_name"]}</div>',
            '<div class="grid">',
            f'<div>货物体积(m³): {_fmt2(cr["volume_m3"])}</div>',
            f'<div>体积利用率: {cr["volume_util"]}%</div>',
            f'<div>货物重量(kg): {_fmt2(cr["weight_kg"])}</div>',
            f'<div>货物和集装箱共重(kg): {_fmt2(cr["gross_kg"])}</div>',
            f'<div>货物价格: {_fmt2(cr["price"])}</div>',
            "</div>",
            f'<div class="kv" style="margin-top:8px">'
            f'长度公差(mm): {_fmt2(cr["tol_length_mm"])}　'
            f'宽度公差(mm): {_fmt2(cr["tol_width_mm"])}　'
            f'高度公差(mm): {_fmt2(cr["tol_height_mm"])}</div>',
            f'<h3>共装上货物(PCS) {cr["total_pcs"]}</h3>',
            _html_table(_item_table_rows(cr["items"])),
        ]
    if d["unpacked_count"]:
        parts.append(f'<p style="color:#b91c1c;font-size:13px">未装入货物: {d["unpacked_count"]} 件</p>')
    parts.append("</body></html>")
    return "".join(parts)


def _html_table(rows: list[list[str]]) -> str:
    if len(rows) <= 1:
        return "<p>（无）</p>"
    head, *body = rows
    th = "".join(f"<th>{c}</th>" for c in head)
    trs = "".join(
        "<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>"
        for r in body
    )
    return f"<table><thead><tr>{th}</tr></thead><tbody>{trs}</tbody></table>"


def build_excel_report(result: dict) -> bytes:
    d = build_report_data(result)
    wb = Workbook()
    ws = wb.active
    ws.title = "报表"

    def row(*cells):
        ws.append(cells)

    row("报表")
    row()
    c = ws["A1"]
    c.font = _TITLE
    row("总体装柜情况")
    ws["A" + str(ws.max_row)].font = _BOLD
    row("总集装箱数", d["total_containers"])
    row("货物总数(PCS)", d["total_pcs"])
    row("货物总体积(m³)", round(d["total_volume_m3"], 2))
    row("货物总重量(kg)", round(d["total_weight_kg"], 2))
    row("货物总价格", round(d["total_price"], 2))
    row()
    row(f'共装上货物(PCS) {d["total_pcs"]}')
    ws["A" + str(ws.max_row)].font = _BOLD
    for r in _item_table_rows(d["items"]):
        row(*r)
    if ws.max_row >= 9:
        for cell in ws[ws.max_row - len(d["items"])]:
            cell.font = _BOLD

    for cr in d["containers"]:
        row()
        row(f'集装箱 {cr["index"]}')
        ws["A" + str(ws.max_row)].font = _BOLD
        row("集装箱", cr["type_name"])
        row("货物体积(m³)", round(cr["volume_m3"], 2), "体积利用率", f'{cr["volume_util"]}%')
        row("货物重量(kg)", round(cr["weight_kg"], 2),
            "货物和集装箱共重(kg)", round(cr["gross_kg"], 2))
        row("货物价格", round(cr["price"], 2))
        row("长度公差(mm)", round(cr["tol_length_mm"], 2),
            "宽度公差(mm)", round(cr["tol_width_mm"], 2),
            "高度公差(mm)", round(cr["tol_height_mm"], 2))
        row(f'共装上货物(PCS) {cr["total_pcs"]}')
        ws["A" + str(ws.max_row)].font = _BOLD
        for r in _item_table_rows(cr["items"]):
            row(*r)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14

    # 明细表(保留)
    for c in result.get("containers", []):
        if not c.get("placements"):
            continue
        title = c["name"][:28] + "明细"
        s = wb.create_sheet(title=title[:31])
        s.append(["货物", "装入序号", "X(cm)", "Y(cm)", "Z(cm)", "长", "宽", "高"])
        for cell in s[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for p in sorted(c["placements"], key=lambda x: x.get("seq", 0)):
            s.append([
                p["base_name"], p.get("seq", ""),
                round(p["x"], 1), round(p["y"], 1), round(p["z"], 1),
                p["dx"], p["dy"], p["dz"],
            ])

    if result.get("unpacked"):
        s = wb.create_sheet(title="未装入")
        s.append(["货物", "长", "宽", "高", "重量"])
        for cell in s[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for u in result["unpacked"]:
            s.append([u["name"], u["length"], u["width"], u["height"], u["weight"]])

    return _save(wb)


def build_pdf_report(result: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    font = "STSong-Light"
    d = build_report_data(result)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                            leftMargin=20 * mm, rightMargin=20 * mm)
    title = ParagraphStyle("title", fontName=font, fontSize=16, leading=20)
    h2 = ParagraphStyle("h2", fontName=font, fontSize=12, leading=16, spaceBefore=10, spaceAfter=6)
    h3 = ParagraphStyle("h3", fontName=font, fontSize=10, leading=14, spaceBefore=6, spaceAfter=4)
    body = ParagraphStyle("body", fontName=font, fontSize=10, leading=14)

    story = [Paragraph("报表", title), Spacer(1, 6)]

    def kv(label, value):
        story.append(Paragraph(f"<b>{label}</b>　{value}", body))

    def add_item_table(items):
        data = _item_table_rows(items)
        t = Table(data, colWidths=[80, 80, 70, 70], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), font),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 8))

    story.append(Paragraph("总体装柜情况", h2))
    kv("总集装箱数", str(d["total_containers"]))
    kv("货物总数(PCS)", str(d["total_pcs"]))
    kv("货物总体积(m³)", _fmt2(d["total_volume_m3"]))
    kv("货物总重量(kg)", _fmt2(d["total_weight_kg"]))
    kv("货物总价格", _fmt2(d["total_price"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f'共装上货物(PCS) {d["total_pcs"]}', h3))
    add_item_table(d["items"])

    for cr in d["containers"]:
        story.append(Spacer(1, 6))
        story.append(Paragraph(f'集装箱 {cr["index"]}', h2))
        kv("集装箱", cr["type_name"])
        kv("货物体积(m³)", f'{_fmt2(cr["volume_m3"])}　　体积利用率: {cr["volume_util"]}%')
        kv("货物重量(kg)", f'{_fmt2(cr["weight_kg"])}　　货物和集装箱共重(kg): {_fmt2(cr["gross_kg"])}')
        kv("货物价格", _fmt2(cr["price"]))
        kv("长度公差(mm)", f'{_fmt2(cr["tol_length_mm"])}　　宽度公差(mm): {_fmt2(cr["tol_width_mm"])}　　高度公差(mm): {_fmt2(cr["tol_height_mm"])}')
        story.append(Paragraph(f'共装上货物(PCS) {cr["total_pcs"]}', h3))
        add_item_table(cr["items"])

    if d["unpacked_count"]:
        story.append(Paragraph(f'未装入货物: {d["unpacked_count"]} 件', body))

    doc.build(story)
    return buf.getvalue()


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
